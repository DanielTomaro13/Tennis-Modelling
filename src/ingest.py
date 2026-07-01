"""Ingest stage — two public sources, each doing the job it's best at.

1. **Match Charting Project** (charting-*.csv): per-match serve/return
   aggregates for the serve profiles. Volunteer-charted subset — the only
   public source of serve stats now that Sackmann's tour-level repos are gone.
2. **tennis-data.co.uk** (per-season .xlsx): EVERY tour-level match — winner,
   loser, surface, best-of, round and closing odds — for Elo, player records
   and the backtest's market baseline. This replaces deriving winners from the
   charted point-by-point files, which covered ~10% of the tour (median Elo
   history was 2 matches per player).

Winner/loser names arrive as "Sinner J." and are mapped to the charting
project's full names ("Jannik Sinner") so profiles, Elo and fixtures share one
name space; unmapped names keep the raw form (they still strengthen the Elo
graph even if never queried).

Run:  python -m src.ingest
"""
from __future__ import annotations

import csv
import datetime
import io
import os
import re
import sys
import unicodedata

from . import util


def _raw_path(cfg: dict, name: str) -> str:
    return util.abspath(os.path.join(cfg["data"]["raw_dir"], name))


# --------------------------------------------------------------------------- #
# Match Charting Project (serve/return profiles)
# --------------------------------------------------------------------------- #
def download_core(cfg: dict) -> None:
    """Download charting matches + Overview CSVs, then the tennis-data files."""
    util.ensure_dir(util.abspath(cfg["data"]["raw_dir"]))
    base = cfg["data"]["base_url"]
    for tour in cfg["tours"]:
        files = cfg["data"]["files"][tour]
        for key in ("matches", "overview"):
            name = files[key]
            dest = _raw_path(cfg, name)
            util.log(f"ingest: downloading {tour}/{key} -> {name}")
            data = util.http_get(f"{base}/{name}")
            with open(dest, "wb") as fh:
                fh.write(data)
            util.log(f"ingest: wrote {dest} ({len(data):,} bytes)")
    download_results(cfg)


def load_matches(cfg: dict, tour: str) -> dict[str, dict]:
    """Return {match_id: {players, surface, date, best_of, tournament, round}}."""
    path = _raw_path(cfg, cfg["data"]["files"][tour]["matches"])
    start = int(cfg["data"]["start_date"])
    out: dict[str, dict] = {}
    with open(path, "r", encoding="utf-8", newline="") as fh:
        for row in csv.DictReader(fh):
            mid = row.get("match_id", "").strip()
            if not mid:
                continue
            try:
                date = int((row.get("Date") or "0").strip() or 0)
            except ValueError:
                date = 0
            if date and date < start:
                continue
            out[mid] = {
                "match_id": mid,
                "tour": tour,
                "p1": (row.get("Player 1") or "").strip(),
                "p2": (row.get("Player 2") or "").strip(),
                "surface": normalize_surface(row.get("Surface")),
                "date": date,
                "best_of": _to_int(row.get("Best of"), 3),
                "tournament": (row.get("Tournament") or "").strip(),
                "round": (row.get("Round") or "").strip(),
            }
    return out


def iter_overview(cfg: dict, tour: str):
    """Yield per-player Total rows from the Overview stats file."""
    path = _raw_path(cfg, cfg["data"]["files"][tour]["overview"])
    with open(path, "r", encoding="utf-8", newline="") as fh:
        for row in csv.DictReader(fh):
            if (row.get("set") or "").strip() != "Total":
                continue
            yield row


# --------------------------------------------------------------------------- #
# tennis-data.co.uk (full tour results + closing odds)
# --------------------------------------------------------------------------- #
def _td_file(tour: str, year: int) -> str:
    return f"td_{tour}_{year}.xlsx"


def _td_url(cfg: dict, tour: str, year: int) -> str:
    base = cfg["data"]["tennis_data_base"].rstrip("/")
    sub = f"{year}" if tour == "atp" else f"{year}w"
    return f"{base}/{sub}/{year}.xlsx"


def download_results(cfg: dict) -> None:
    """Fetch per-season tennis-data workbooks. Past seasons cached; the current
    season is always refreshed (it grows weekly)."""
    this_year = datetime.date.today().year
    start = int(cfg["data"].get("start_year", 2010))
    for tour in cfg["tours"]:
        for year in range(start, this_year + 1):
            dest = _raw_path(cfg, _td_file(tour, year))
            fresh = year >= this_year - (1 if datetime.date.today().month <= 2 else 0)
            if os.path.exists(dest) and not fresh:
                continue
            try:
                data = util.http_get(_td_url(cfg, tour, year))
            except Exception as exc:
                if not os.path.exists(dest):
                    util.log(f"ingest: tennis-data {tour} {year} unavailable ({exc})")
                continue
            with open(dest, "wb") as fh:
                fh.write(data)
            util.log(f"ingest: wrote {_td_file(tour, year)} ({len(data):,} bytes)")


def _strip_accents(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def _name_keys_full(full: str) -> list[str]:
    """Keys for a charting full name: '<surname tokens>|<first initial>' for
    every possible split (handles 'Alex De Minaur', 'Juan Martin del Potro')."""
    toks = _strip_accents(full).lower().replace("-", " ").split()
    if len(toks) < 2:
        return []
    return [f"{' '.join(toks[i:])}|{toks[0][0]}" for i in range(1, len(toks))]


def _name_key_td(name: str) -> str | None:
    """Key for a tennis-data name: 'Sinner J.' / 'De Minaur A.' / 'Del Potro J.M.'."""
    s = _strip_accents((name or "").strip()).lower().replace("-", " ")
    m = re.match(r"^(.*?)\s+((?:[a-z]\.\s*)+)$", s)
    if not m:
        return None
    surname = " ".join(m.group(1).split())
    first_initial = m.group(2).strip()[0]
    return f"{surname}|{first_initial}"


def build_name_map(cfg: dict, tour: str) -> dict[str, str]:
    """tennis-data key -> charting full name (collisions resolved by charted
    volume, so the well-known player wins)."""
    counts: dict[str, float] = {}
    for row in iter_overview(cfg, tour):
        p = (row.get("player") or "").strip()
        try:
            counts[p] = counts.get(p, 0.0) + float(row.get("serve_pts") or 0)
        except (TypeError, ValueError):
            counts[p] = counts.get(p, 0.0)
    out: dict[str, str] = {}
    strength: dict[str, float] = {}
    for full, vol in counts.items():
        for key in _name_keys_full(full):
            if key not in out or vol > strength[key]:
                out[key] = full
                strength[key] = vol
    return out


def _cell(row: dict, *names):
    for n in names:
        if n in row and row[n] not in (None, ""):
            return row[n]
    return None


def load_results(cfg: dict, tour: str) -> list[dict]:
    """All tour matches from the cached tennis-data workbooks, chronological,
    names translated to charting full names where possible."""
    import openpyxl

    name_map = build_name_map(cfg, tour)
    this_year = datetime.date.today().year
    start = int(cfg["data"].get("start_year", 2010))
    out: list[dict] = []
    for year in range(start, this_year + 1):
        path = _raw_path(cfg, _td_file(tour, year))
        if not os.path.exists(path):
            continue
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:  # legacy .xls or an HTML error page — skip season
            util.log(f"ingest: {_td_file(tour, year)} unreadable ({exc}) — skipping")
            continue
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = [str(h) if h is not None else "" for h in next(rows)]
        for vals in rows:
            row = dict(zip(header, vals))
            w_raw, l_raw = _cell(row, "Winner"), _cell(row, "Loser")
            if not w_raw or not l_raw:
                continue
            d = _cell(row, "Date")
            if isinstance(d, datetime.datetime):
                date = d.year * 10000 + d.month * 100 + d.day
            else:
                try:  # DD/MM/YYYY string in some sheets
                    dd, mm, yy = str(d).split("/")
                    date = int(yy) * 10000 + int(mm) * 100 + int(dd)
                except (AttributeError, ValueError):
                    continue
            wk, lk = _name_key_td(str(w_raw)), _name_key_td(str(l_raw))
            odds_w = _num(_cell(row, "AvgW", "B365W", "PSW"))
            odds_l = _num(_cell(row, "AvgL", "B365L", "PSL"))
            out.append({
                "winner": name_map.get(wk, str(w_raw).strip()),
                "loser": name_map.get(lk, str(l_raw).strip()),
                "surface": normalize_surface(_cell(row, "Surface")),
                "date": date,
                "best_of": _to_int(_cell(row, "Best of"), 3),
                "tournament": str(_cell(row, "Tournament") or "").strip(),
                "round": str(_cell(row, "Round") or "").strip(),
                "odds_w": odds_w, "odds_l": odds_l,
            })
        wb.close()
    out.sort(key=lambda r: r["date"])
    return out


def derive_winners(cfg: dict) -> None:
    """Write results-{tour}.csv from the tennis-data workbooks."""
    proc = util.ensure_dir(util.abspath(cfg["data"]["processed_dir"]))
    for tour in cfg["tours"]:
        results = load_results(cfg, tour)
        out = util.abspath(os.path.join(proc, f"results-{tour}.csv"))
        with open(out, "w", encoding="utf-8", newline="") as fh:
            w = csv.writer(fh)
            w.writerow(["winner", "loser", "surface", "date", "best_of",
                        "tournament", "round", "odds_w", "odds_l"])
            for m in results:
                w.writerow([m["winner"], m["loser"], m["surface"], m["date"], m["best_of"],
                            m["tournament"], m["round"], m["odds_w"] or "", m["odds_l"] or ""])
        util.log(f"ingest: wrote {out} ({len(results):,} results)")


# --------------------------------------------------------------------------- #
def normalize_surface(value) -> str:
    s = (str(value) if value is not None else "").strip().lower()
    if s.startswith("hard"):
        return "Hard"
    if s.startswith("clay"):
        return "Clay"
    if s.startswith("grass"):
        return "Grass"
    if s.startswith("carpet"):
        return "Carpet"
    return "Hard"  # sensible default for the rare blank


def _to_int(value, default: int) -> int:
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return default


def _num(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def main(argv: list[str]) -> int:
    cfg = util.load_config()
    download_core(cfg)
    derive_winners(cfg)
    for tour in cfg["tours"]:
        matches = load_matches(cfg, tour)
        n_over = sum(1 for _ in iter_overview(cfg, tour))
        util.log(f"ingest: {tour}: {len(matches):,} charted matches, {n_over:,} overview rows")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
